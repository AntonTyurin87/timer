from openpyxl import load_workbook
from datetime import datetime

# Дата и предустановки
def date_position():
    global date_place, file_page, file_data, time_data
    time_data = '/home/anton/repositories/timer/time_stat.xlsx'
    file_data = load_workbook(time_data)
    file_page = file_data['data']
    date_place = len(file_page['A'])
    if file_page['A' + str(date_place)].value != datetime.today().strftime('%Y-%m-%d'):
        date_place = len(file_page['A']) + 1


# Запись даты и необходимого времени
def date_and_time_wrote(time_work_plane, time_studies_plane, time_rest_plane):
    time_data = '/home/anton/repositories/timer/time_stat.xlsx'

    file_data = load_workbook(time_data)
    file_page = file_data['data']

    work_plane_minutes = time_work_plane[0] * 60 + time_work_plane[1]
    studies_plane_minutes = time_studies_plane[0] * 60 + time_studies_plane[1]
    rest_plane_minutes = time_rest_plane[0] * 60 + time_rest_plane[1]
    
    date_place = len(file_page['A'])

    if file_page['A' + str(date_place)].value != datetime.today().strftime('%Y-%m-%d'):
        file_page['A' + str(date_place + 1)].value = datetime.today().strftime('%Y-%m-%d')
        file_page['B' + str(date_place + 1)] = work_plane_minutes
        file_page['D' + str(date_place + 1)] = studies_plane_minutes
        file_page['F' + str(date_place + 1)] = rest_plane_minutes

    elif file_page['A' + str(date_place)].value == datetime.today().strftime('%Y-%m-%d'):
        file_page['B' + str(date_place)] = work_plane_minutes
        file_page['D' + str(date_place)] = studies_plane_minutes
        file_page['F' + str(date_place)] = rest_plane_minutes

    file_data.save(time_data)


# Запись количества прошедших процентов
def persent_pass(persent, action_tipe):
    if action_tipe == 'work':
        file_page['C' + str(date_place)] = persent
        file_data.save(time_data)

    elif action_tipe == 'studies':
        file_page['E' + str(date_place)] = persent
        file_data.save(time_data)

    elif action_tipe == 'rest':
        file_page['G' + str(date_place)] = persent
        file_data.save(time_data)
        

# Статистика за 5 дней
def five_days_time():
    date_position()
    if len(file_page['A']) < 6:
        return 'no_data'
    else:
        summ_minutes_work = 0
        summ_minutes_studies = 0
        summ_minutes_rest = 0

        persent_work = 0
        persent_studies = 0
        persent_rest = 0

        index_string = len(file_page['A'])
        for i in range(1, 6):
            summ_minutes_work += int(file_page['B' + str(index_string - i)].value)
            summ_minutes_studies += int(file_page['D' + str(index_string - i)].value)
            summ_minutes_rest += int(file_page['F' + str(index_string - i)].value)

            persent_work += file_page['C' + str(index_string - i)].value
            persent_studies += file_page['E' + str(index_string - i)].value
            persent_rest += file_page['G' + str(index_string - i)].value
    
        return [(summ_minutes_work, persent_work/5),
                (summ_minutes_studies, persent_studies/5),
                (summ_minutes_rest, persent_rest/5)]
        

# Статистика за 21 день
def twentyone_days_time():
    date_position()
    if len(file_page['A']) < 22:
        return 'no_data'
    else:
        summ_minutes_work = 0
        summ_minutes_studies = 0
        summ_minutes_rest = 0

        persent_work = 0
        persent_studies = 0
        persent_rest = 0

        index_string = len(file_page['A'])
        for i in range(1, 22):
            summ_minutes_work += int(file_page['B' + str(index_string - i)].value)
            summ_minutes_studies += int(file_page['D' + str(index_string - i)].value)
            summ_minutes_rest += int(file_page['F' + str(index_string - i)].value)

            persent_work += file_page['C' + str(index_string - i)].value
            persent_studies += file_page['E' + str(index_string - i)].value
            persent_rest += file_page['G' + str(index_string - i)].value
    
        return [(summ_minutes_work, persent_work/21),
                (summ_minutes_studies, persent_studies/21),
                (summ_minutes_rest, persent_rest/21)]
        
        
