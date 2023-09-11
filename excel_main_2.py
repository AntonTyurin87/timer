from openpyxl import load_workbook
from datetime import datetime


def check_date_position():
    global file_page, time_data
    time_data  = '/home/anton/repositories/timer/time_stat_2.xlsx'
    file_data = load_workbook(time_data)
    date_now = datetime.today().strftime('%Y-%m-%d')
    file_page = file_data['data']
    date_string = len(file_page['A'])
    if file_page['A' + str(date_string)].value != date_now:
        date_string = len(file_page['A']) + 1
        file_page['A' + str(date_string)].value = date_now
        file_data.save(time_data)
    else:
        date_string = len(file_page['A'])

    return date_string


def set_time_plane(date_string, button_number, scheduled_time):
    column = ['A', 'B', 'D', 'F']
    column_s = ['A', 'C', 'E', 'G']
    time_data  = '/home/anton/repositories/timer/time_stat_2.xlsx'
    file_data = load_workbook(time_data)
    file_page = file_data['data']
    #print(file_page[column[bottun_number] + str(date_string)].value)
    #print(scheduled_time.get(1)[1])
    if file_page[column[button_number] + str(date_string)].value  == None:
        for i in range(1, len(column)):
            minutes = scheduled_time.get(i)[0] * 60 + scheduled_time.get(i)[1]
            file_page[column[i] + str(date_string)].value = minutes
            file_page[column_s[i] + str(date_string)].value = 0
            #print(file_page[column[i] + str(date_string)].value)
        file_data.save(time_data)

#print(set_time_plane(2, 1, {1 :[2, 30], 2: [1, 26], 3: [3, 56]}))


def set_time_pass(date_string, button_number, time_count):
    column_s = ['A', 'C', 'E', 'G']
    time_data  = '/home/anton/repositories/timer/time_stat_2.xlsx'
    file_data = load_workbook(time_data)
    file_page = file_data['data']
    file_page[column_s[button_number] + str(date_string)].value += int(time_count)
    file_data.save(time_data)

    return file_page[column_s[button_number] + str(date_string)].value



    

            



'''
def date_position():
    global date_place, file_page, file_data, time_data
    time_data = '/home/anton/repositories/timer/time_stat.xlsx'
    file_data = load_workbook(time_data)
    file_page = file_data['data']
    date_place = len(file_page['A'])
    if file_page['A' + str(date_place)].value != datetime.today().strftime('%Y-%m-%d'):
        date_place = len(file_page['A']) + 1
'''